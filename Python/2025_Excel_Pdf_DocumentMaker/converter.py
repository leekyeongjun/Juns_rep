import os
from openpyxl import workbook
from openpyxl import load_workbook
from pptx import Presentation
from pptxtopdf import convert
import fitz # pip install PyMuPDF

class Bot:
    def deleteAllFiles(self, filepath) :
        if os.path.exists(filepath):
            for file in os.scandir(filepath):
                os.remove(file.path)

    def __init__(self):
        self.xl_fileName = None
        self.xl_workbook = None
        self.xl_sheetNames = []
        self.xl_sheetScopes = [] # 시작점과 마지막 점 두개의 배열들의 배열

        self.pp_fileName = None
        self.pp_workslide = None
        self.targetDirectory = None

        self.allData = []

        self.replaces = {}
        self.dataType = 0 #0은 미정, 1은 가로형, 2는 세로형
    
    def SetWorkBook(self, _fileName):
        try :
            self.xl_fileName = _fileName
            self.xl_workbook = load_workbook(filename= self.xl_fileName)
        except :
            if (_fileName.endswith('.xlsx') or _fileName.endswith('xlsm')) == False :
                print("확장자 에러")
                return 1
            else :
                print("알 수 없는 오류")
                return 2
        
        print(self.xl_workbook)
        return 0

    def SetPPTX(self, _fileName):
        if  _fileName.endswith('.pptx'):
            self.deleteAllFiles(".\\thumbs\\")
            self.pp_fileName = _fileName
            self.pp_workslide = Presentation(self.pp_fileName)
            print(self.pp_fileName)
            filename_with_extension = os.path.basename(self.pp_fileName)
            print(filename_with_extension)
            filename_without_extension = os.path.splitext(filename_with_extension)[0]
            print(filename_without_extension)

            pdfpath = "./thumbs/"

            convert(_fileName, pdfpath)
            thumbfile = fitz.open(pdfpath+filename_without_extension+".pdf")

            img = thumbfile[0].get_pixmap()

            self.thumbpath = f"./thumbs/{filename_without_extension}.png"
            img.save(self.thumbpath)

            return self.thumbpath
        else :
            return 1
        
        
    def SetAllData(self, sheetIds):
        # xl_fileName과 dataType을 기반으로 allData 설정
        for sheetId in sheetIds :
            try :
                sheetName = self.xl_workbook.sheetnames[sheetId]
                self.xl_sheetNames.append(sheetName)
                worksheet = self.xl_workbook[sheetName]
                if self.dataType == 1 :
                    self.allData.append(worksheet.values)
                elif self.dataType == 2:
                    tmpData = []
                    for col in worksheet.iter_cols(values_only=True):
                        tmpData.append(col)

                    self.allData.append(tmpData)
                else :
                    print("데이터 타입 미설정")
                    self.allData.clear()
                    return 1
            except :
                print("데이터 파싱 중 오류 발생")
                self.allData.clear()
                return 2
            
        return 0

    def ShowAllData(self):
        for sheetdata in self.allData:
            for value in sheetdata :
                print(value)

    def ShowData(self, sheetId):
        sheetName = self.xl_workbook.sheetnames[sheetId]
        print(sheetName)
        for value in self.allData[sheetId]:
            print(value)
    
## DEBUG AREA
"""
bot = Bot()

bot.SetWorkBook("./test.xlsx")

bot.dataType = 2

print(bot.xl_workbook.sheetnames)

bot.SetAllData([0,1,2])

for i in range(0,3):
    bot.ShowData(i)
"""
